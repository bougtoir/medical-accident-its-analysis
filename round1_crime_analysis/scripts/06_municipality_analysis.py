"""
06_municipality_analysis.py
Municipality-level crime analysis for Round1 Japan (Tokyo focus).

Uses 警視庁 町丁別認知件数 data (2009-2025) to compare:
  - Round1 municipalities vs non-Round1 municipalities within Tokyo
  - Pre/post opening event study for municipalities with treatment windows

Data sources:
  - Tokyo crime: 警視庁 区市町村の町丁別、罪種別及び手口別認知件数
    XLS: H21-H25 (2009-2013), XLSX: H28 (2016)
    CSV: H29-H30 (2017-2018), R2-R7 (2020-2025)
  - Osaka crime: 大阪府警 刑法犯罪種及び手口別発生市区町村別認知件数 (R7 only)
  - Round1 stores: round1_japan_stores.json
"""

import os
import json
import csv
import warnings

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from scipy import stats
import statsmodels.formula.api as smf

warnings.filterwarnings("ignore", category=FutureWarning)

BASE_DIR = os.path.join(os.path.dirname(__file__), "..")
DATA_DIR = os.path.join(BASE_DIR, "data")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
FIG_DIR = os.path.join(BASE_DIR, "figures")
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(FIG_DIR, exist_ok=True)

C_R1 = "#E63946"
C_CTRL = "#457B9D"
C_ACCENT = "#2A9D8F"
C_DARK = "#1D3557"


# ═══════════════════════════════════════════════════════════════════
# Data loading: Tokyo crime CSVs
# ═══════════════════════════════════════════════════════════════════
CSV_YEARS = {
    "H29": 2017, "H30": 2018,
    "R2": 2020, "R3": 2021, "R4": 2022, "R5": 2023, "R6": 2024, "R7": 2025,
}


def parse_tokyo_csv(path, year):
    """Parse a Tokyo crime CSV into municipality-level aggregates."""
    municipality_totals = {}
    with open(path, "r", encoding="cp932") as f:
        reader = csv.reader(f)
        header = next(reader)
        for row in reader:
            loc = row[0].strip()
            if "計" in loc or "不詳" in loc or "他県" in loc or "海外" in loc:
                continue
            try:
                total = int(row[1])
            except (ValueError, IndexError):
                continue
            violent = int(row[2]) if len(row) > 2 and row[2].strip() else 0
            rough = int(row[5]) if len(row) > 5 and row[5].strip() else 0
            burglary = int(row[11]) if len(row) > 11 and row[11].strip() else 0
            theft_non = int(row[20]) if len(row) > 20 and row[20].strip() else 0
            other = int(row[32]) if len(row) > 32 and row[32].strip() else 0

            municipality = None
            for suffix in ["区", "市", "町", "村"]:
                idx = loc.find(suffix)
                if idx >= 0:
                    municipality = loc[: idx + 1]
                    break

            if municipality:
                if municipality not in municipality_totals:
                    municipality_totals[municipality] = {
                        "total": 0, "violent": 0, "rough": 0,
                        "burglary": 0, "theft_non": 0, "other": 0,
                    }
                m = municipality_totals[municipality]
                m["total"] += total
                m["violent"] += violent
                m["rough"] += rough
                m["burglary"] += burglary
                m["theft_non"] += theft_non
                m["other"] += other
    return municipality_totals


def parse_tokyo_xls(path, year):
    """Parse a Tokyo crime XLS (H21-H25) into municipality-level aggregates."""
    import xlrd
    wb = xlrd.open_workbook(path)
    municipality_totals = {}
    skip_sheets = {"最終", "合計", "総計"}
    for sname in wb.sheet_names():
        if sname in skip_sheets:
            continue
        ws = wb.sheet_by_name(sname)
        totals = {"total": 0, "violent": 0, "rough": 0,
                  "burglary": 0, "theft_non": 0, "other": 0}
        for r in range(7, ws.nrows):
            loc = str(ws.cell_value(r, 0)).strip()
            if "計" in loc or "不詳" in loc or not loc:
                continue
            def cell_int(c):
                v = ws.cell_value(r, c)
                return int(v) if isinstance(v, (int, float)) and v > 0 else 0
            totals["total"] += cell_int(1)
            totals["violent"] += cell_int(2)
            totals["rough"] += cell_int(5)
            totals["burglary"] += cell_int(11)
            totals["theft_non"] += cell_int(20)
            totals["other"] += cell_int(32)
        municipality_totals[sname] = totals
    return municipality_totals


def parse_tokyo_xlsx(path, year):
    """Parse a Tokyo crime XLSX (H28) into municipality-level aggregates."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    municipality_totals = {}
    skip_sheets = {"最終", "合計", "総計"}
    for sname in wb.sheetnames:
        if sname in skip_sheets:
            continue
        ws = wb[sname]
        totals = {"total": 0, "violent": 0, "rough": 0,
                  "burglary": 0, "theft_non": 0, "other": 0}
        for row in ws.iter_rows(min_row=8, values_only=True):
            loc = str(row[0]).strip() if row[0] else ""
            if "計" in loc or "不詳" in loc or not loc:
                continue
            def cell_int(idx):
                v = row[idx] if idx < len(row) else 0
                return int(v) if isinstance(v, (int, float)) and v and v > 0 else 0
            totals["total"] += cell_int(1)
            totals["violent"] += cell_int(2)
            totals["rough"] += cell_int(5)
            totals["burglary"] += cell_int(11)
            totals["theft_non"] += cell_int(20)
            totals["other"] += cell_int(32)
        municipality_totals[sname] = totals
    return municipality_totals


def load_all_tokyo_crime():
    """Load all available Tokyo municipality crime data."""
    tokyo_dir = os.path.join(DATA_DIR, "tokyo_crime")
    all_data = {}

    # XLS files (H21-H25 = 2009-2013)
    xls_years = {"H21": 2009, "H22": 2010, "H23": 2011, "H24": 2012, "H25": 2013}
    for fname, year in xls_years.items():
        path = os.path.join(tokyo_dir, f"{fname}.xls")
        if os.path.exists(path):
            all_data[year] = parse_tokyo_xls(path, year)

    # XLSX (H28 = 2016)
    xlsx_path = os.path.join(tokyo_dir, "H28.xlsx")
    if os.path.exists(xlsx_path):
        all_data[2016] = parse_tokyo_xlsx(xlsx_path, 2016)

    # CSV files
    for fname, year in CSV_YEARS.items():
        path = os.path.join(tokyo_dir, f"{fname}.csv")
        if os.path.exists(path) and os.path.getsize(path) > 50000:
            all_data[year] = parse_tokyo_csv(path, year)

    return all_data


def build_tokyo_panel(all_data, stores):
    """Build panel dataframe: municipality × year with R1 treatment info."""
    tokyo_stores = [s for s in stores if s["prefecture"] == "東京都"]
    r1_municipalities = {}
    for s in tokyo_stores:
        city = s["city"]
        open_year = int(s["open_date"][:4])
        for suffix in ["区", "市", "町", "村"]:
            idx = city.find(suffix)
            if idx >= 0:
                muni = city[: idx + 1]
                if muni not in r1_municipalities or open_year < r1_municipalities[muni]:
                    r1_municipalities[muni] = open_year
                break

    rows = []
    for year, muni_data in sorted(all_data.items()):
        for muni, crimes in muni_data.items():
            first_r1 = r1_municipalities.get(muni, np.nan)
            rows.append({
                "year": year,
                "municipality": muni,
                "total_crimes": crimes["total"],
                "violent": crimes["violent"],
                "rough": crimes["rough"],
                "burglary": crimes["burglary"],
                "theft_non": crimes["theft_non"],
                "other": crimes["other"],
                "has_r1": 1 if muni in r1_municipalities else 0,
                "first_r1_year": first_r1,
                "post_r1": 1 if not np.isnan(first_r1) and year >= first_r1 else 0,
            })

    df = pd.DataFrame(rows)
    return df, r1_municipalities


# ═══════════════════════════════════════════════════════════════════
# Analysis functions
# ═══════════════════════════════════════════════════════════════════
def run_tokyo_did(df):
    """DiD regression: R1 municipalities vs non-R1 within Tokyo."""
    print(f"\n{'='*60}")
    print("TOKYO MUNICIPALITY-LEVEL DiD")
    print(f"{'='*60}")
    print(f"Years: {sorted(df['year'].unique())}")
    print(f"R1 municipalities: {df[df['has_r1']==1]['municipality'].nunique()}")
    print(f"Control municipalities: {df[df['has_r1']==0]['municipality'].nunique()}")

    results = {}
    for title, col in [
        ("Total Crimes", "total_crimes"),
        ("Violent Crime", "violent"),
        ("Assault/Battery", "rough"),
        ("Burglary", "burglary"),
        ("Non-intrusion Theft", "theft_non"),
    ]:
        reg_data = df[["municipality", "year", col, "post_r1"]].dropna()
        reg_data = reg_data.rename(columns={col: "crime"})
        try:
            model = smf.ols(
                "crime ~ post_r1 + C(municipality) + C(year)", data=reg_data
            ).fit(cov_type="cluster", cov_kwds={"groups": reg_data["municipality"]})
            coef = model.params["post_r1"]
            se = model.bse["post_r1"]
            pval = model.pvalues["post_r1"]
            ci_lo, ci_hi = model.conf_int().loc["post_r1"]
            sig = "***" if pval < 0.001 else "**" if pval < 0.01 else "*" if pval < 0.05 else ""

            results[title] = {
                "coefficient": float(round(coef, 1)),
                "std_error": float(round(se, 1)),
                "p_value": float(round(pval, 4)),
                "ci_95": [float(round(ci_lo, 1)), float(round(ci_hi, 1))],
                "significant": bool(pval < 0.05),
                "n_obs": int(len(reg_data)),
            }
            print(f"\n  {title}:")
            print(f"    DiD coef:  {coef:+.1f} {sig}")
            print(f"    SE:        {se:.1f}")
            print(f"    p-value:   {pval:.4f}")
            print(f"    95% CI:    [{ci_lo:.1f}, {ci_hi:.1f}]")
        except Exception as e:
            print(f"\n  {title}: failed — {e}")
            results[title] = {"error": str(e)}

    with open(os.path.join(OUTPUT_DIR, "tokyo_municipality_did.json"), "w") as f:
        json.dump(results, f, indent=2)
    return results


def plot_tokyo_trends(df, r1_municipalities):
    """Crime trends: R1 vs non-R1 municipalities in Tokyo."""
    fig, axes = plt.subplots(1, 2, figsize=(14, 6))

    for idx, (title, col) in enumerate([
        ("Total Crimes", "total_crimes"),
        ("Non-intrusion Theft", "theft_non"),
    ]):
        ax = axes[idx]
        for grp, color, label in [
            (1, C_R1, f"Round1 municipalities (n={df[df['has_r1']==1]['municipality'].nunique()})"),
            (0, C_CTRL, f"Non-R1 municipalities (n={df[df['has_r1']==0]['municipality'].nunique()})"),
        ]:
            grp_data = df[df["has_r1"] == grp].groupby("year")[col].mean()
            ax.plot(grp_data.index, grp_data.values, color=color, lw=2,
                    marker="o", markersize=4, label=label)
        ax.set_xlabel("Year", fontsize=11)
        ax.set_ylabel(f"Mean {title} per Municipality", fontsize=11)
        ax.set_title(title, fontsize=12, fontweight="bold")
        ax.legend(fontsize=8)
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, p: f"{x:,.0f}"))

    plt.suptitle(
        "Tokyo: Round1 vs Non-R1 Municipality Crime Trends",
        fontsize=13, fontweight="bold",
    )
    plt.tight_layout()
    fig.savefig(os.path.join(FIG_DIR, "fig_jp8_tokyo_municipality_trends.png"), dpi=200)
    plt.close(fig)
    print("[Fig JP8] Tokyo municipality trends saved.")


def plot_tokyo_forest(results):
    """Forest plot for Tokyo municipality DiD."""
    labels, coefs, cis_lo, cis_hi, colors = [], [], [], [], []
    for crime, res in results.items():
        if "error" in res:
            continue
        labels.append(crime)
        coefs.append(res["coefficient"])
        cis_lo.append(res["ci_95"][0])
        cis_hi.append(res["ci_95"][1])
        colors.append(C_R1 if res["significant"] else C_CTRL)

    if not labels:
        return

    y_pos = np.arange(len(labels))
    err_lo = [c - lo for c, lo in zip(coefs, cis_lo)]
    err_hi = [hi - c for c, hi in zip(coefs, cis_hi)]

    fig, ax = plt.subplots(figsize=(10, 5))
    ax.errorbar(coefs, y_pos, xerr=[err_lo, err_hi],
                fmt="o", markersize=8, capsize=5,
                color=C_DARK, ecolor="gray", elinewidth=1.5)
    for i, (c, col) in enumerate(zip(coefs, colors)):
        ax.plot(c, i, "o", color=col, markersize=8, zorder=5)
    ax.axvline(x=0, color="gray", ls="--", alpha=0.7)
    ax.set_yticks(y_pos)
    ax.set_yticklabels(labels, fontsize=10)
    ax.set_xlabel("DiD Coefficient (change in crime count)", fontsize=11)
    ax.set_title(
        "Tokyo Municipality DiD: Round1 vs Non-R1 Areas\n"
        "(Red = p<0.05, Blue = not significant)",
        fontsize=12, fontweight="bold",
    )
    ax.invert_yaxis()
    plt.tight_layout()
    fig.savefig(os.path.join(FIG_DIR, "fig_jp9_tokyo_did_forest.png"), dpi=200)
    plt.close(fig)
    print("[Fig JP9] Tokyo municipality DiD forest plot saved.")


MUNI_ROMAJI = {
    "江東区": "Koto-ku",
    "町田市": "Machida-shi",
    "板橋区": "Itabashi-ku",
    "武蔵村山市": "Musashi-Murayama-shi",
    "府中市": "Fuchu-shi",
}


def plot_individual_municipalities(df, r1_municipalities):
    """Per-municipality crime trends for R1 areas with treatment lines."""
    r1_munis = sorted(r1_municipalities.keys())
    n = len(r1_munis)
    if n == 0:
        return

    cols = min(3, n)
    rows_n = (n + cols - 1) // cols
    fig, axes = plt.subplots(rows_n, cols, figsize=(5 * cols, 4 * rows_n))
    if n == 1:
        axes = [axes]
    else:
        axes = axes.flatten()

    for i, muni in enumerate(r1_munis):
        ax = axes[i]
        muni_data = df[df["municipality"] == muni].sort_values("year")
        ax.plot(muni_data["year"], muni_data["total_crimes"],
                color=C_R1, lw=2, marker="o", markersize=4)
        open_year = r1_municipalities[muni]
        ax.axvline(x=open_year, color="gray", ls="--", alpha=0.7,
                   label=f"R1 opened {open_year}")
        romaji = MUNI_ROMAJI.get(muni, muni)
        ax.set_title(f"{romaji} (R1: {open_year})", fontsize=11, fontweight="bold")
        ax.set_xlabel("Year")
        ax.set_ylabel("Total Crimes")
        ax.legend(fontsize=8)
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, p: f"{x:,.0f}"))

    for j in range(i + 1, len(axes)):
        axes[j].set_visible(False)

    plt.suptitle("Crime Trends in Round1 Municipalities (Tokyo)",
                 fontsize=13, fontweight="bold")
    plt.tight_layout()
    fig.savefig(os.path.join(FIG_DIR, "fig_jp10_r1_municipality_detail.png"), dpi=200)
    plt.close(fig)
    print("[Fig JP10] Individual R1 municipality trends saved.")


# Tokyo municipality adjacency (R1 municipalities → neighbors)
TOKYO_MUNI_NEIGHBORS = {
    "江東区": ["中央区", "墨田区", "江戸川区", "品川区", "港区"],
    "町田市": ["八王子市", "多摩市", "相模原市"],
    "板橋区": ["北区", "練馬区", "豊島区"],
    "武蔵村山市": ["東大和市", "西多摩郡瑞穂町", "立川市", "昭島市"],
    "府中市": ["調布市", "小金井市", "国分寺市", "多摩市", "稲城市", "国立市"],
}


def run_neighbor_municipality_did(df, r1_municipalities):
    """DiD comparing R1 municipalities to their adjacent non-R1 neighbors."""
    print(f"\n{'='*60}")
    print("NEIGHBOR MUNICIPALITY DiD (Tokyo)")
    print(f"{'='*60}")

    r1_munis = set(r1_municipalities.keys())
    neighbor_non_r1 = set()
    for muni in r1_munis:
        for nb in TOKYO_MUNI_NEIGHBORS.get(muni, []):
            if nb not in r1_munis:
                neighbor_non_r1.add(nb)

    relevant = r1_munis | neighbor_non_r1
    sub = df[df["municipality"].isin(relevant)].copy()
    print(f"  R1 municipalities: {len(r1_munis)}")
    print(f"  Adjacent non-R1: {len(neighbor_non_r1)}")
    print(f"  Observations: {len(sub)}")

    if len(sub) < 10:
        print("  [SKIP] too few observations")
        return None

    results = {}
    for title, col in [
        ("Total Crimes", "total_crimes"),
        ("Assault/Battery", "rough"),
        ("Burglary", "burglary"),
        ("Non-intrusion Theft", "theft_non"),
    ]:
        reg_data = sub[["municipality", "year", col, "post_r1"]].dropna()
        reg_data = reg_data.rename(columns={col: "crime"})
        try:
            model = smf.ols(
                "crime ~ post_r1 + C(municipality) + C(year)", data=reg_data
            ).fit(cov_type="cluster", cov_kwds={"groups": reg_data["municipality"]})
            coef = model.params["post_r1"]
            pval = model.pvalues["post_r1"]
            sig = "***" if pval < 0.001 else "**" if pval < 0.01 else "*" if pval < 0.05 else ""
            results[title] = {
                "coefficient": float(round(coef, 1)),
                "p_value": float(round(pval, 4)),
            }
            print(f"  {title}: coef={coef:+.1f} {sig} (p={pval:.4f})")
        except Exception as e:
            print(f"  {title}: failed — {e}")

    with open(os.path.join(OUTPUT_DIR, "tokyo_neighbor_municipality_did.json"), "w") as f:
        json.dump(results, f, indent=2)
    return results


def plot_neighbor_municipality_trends(df, r1_municipalities):
    """Plot R1 vs neighbor non-R1 municipality crime trends."""
    r1_munis = set(r1_municipalities.keys())
    neighbor_non_r1 = set()
    for muni in r1_munis:
        for nb in TOKYO_MUNI_NEIGHBORS.get(muni, []):
            if nb not in r1_munis:
                neighbor_non_r1.add(nb)

    relevant = r1_munis | neighbor_non_r1
    sub = df[df["municipality"].isin(relevant)].copy()
    sub["group"] = sub["municipality"].apply(
        lambda m: "Round1" if m in r1_munis else "Neighbor"
    )

    fig, ax = plt.subplots(figsize=(10, 6))
    for grp, color, label in [
        ("Round1", C_R1, f"R1 municipalities (n={len(r1_munis)})"),
        ("Neighbor", C_CTRL, f"Adjacent non-R1 (n={len(neighbor_non_r1)})"),
    ]:
        g = sub[sub["group"] == grp].groupby("year")["total_crimes"].mean()
        ax.plot(g.index, g.values, color=color, lw=2,
                marker="o", markersize=5, label=label)
    ax.set_xlabel("Year", fontsize=11)
    ax.set_ylabel("Mean Total Crimes per Municipality", fontsize=11)
    ax.set_title("Tokyo: Round1 vs Adjacent Non-R1 Municipalities",
                 fontsize=13, fontweight="bold")
    ax.legend(fontsize=10)
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, p: f"{x:,.0f}"))
    plt.tight_layout()
    fig.savefig(os.path.join(FIG_DIR, "fig_jp11_neighbor_municipality_trends.png"), dpi=200)
    plt.close(fig)
    print("[Fig JP11] Neighbor municipality trends saved.")


def analyze_osaka_snapshot(stores):
    """Analyze Osaka R7 municipality data (snapshot comparison)."""
    import openpyxl
    osaka_path = os.path.join(DATA_DIR, "osaka_crime", "hanzaitokei09_r07.xlsx")
    if not os.path.exists(osaka_path):
        print("  [SKIP] Osaka data not found")
        return None

    wb = openpyxl.load_workbook(osaka_path, data_only=True)
    ws = wb.active

    municipalities = {}
    for row in ws.iter_rows(min_row=10, max_row=142, values_only=True):
        muni = row[1]
        total = row[3]
        if muni and total and isinstance(total, (int, float)):
            muni_str = str(muni).strip()
            if "計" not in muni_str and "総数" not in muni_str:
                municipalities[muni_str] = int(total)

    osaka_stores = [s for s in stores if s["prefecture"] == "大阪府"]
    r1_munis = set()
    for s in osaka_stores:
        city = s["city"]
        for muni in municipalities:
            if city.startswith(muni) or muni in city:
                r1_munis.add(muni)
                break

    r1_total = [v for k, v in municipalities.items() if k in r1_munis]
    ctrl_total = [v for k, v in municipalities.items() if k not in r1_munis]

    print(f"\n{'='*60}")
    print("OSAKA MUNICIPALITY SNAPSHOT (R7/2025)")
    print(f"{'='*60}")
    print(f"  R1 municipalities: {len(r1_total)} (matched)")
    print(f"  Non-R1 municipalities: {len(ctrl_total)}")
    if r1_total and ctrl_total:
        print(f"  R1 mean crimes: {np.mean(r1_total):.0f}")
        print(f"  Non-R1 mean: {np.mean(ctrl_total):.0f}")
        t, p = stats.ttest_ind(r1_total, ctrl_total)
        print(f"  t-test: t={t:.3f}, p={p:.4f}")

    return {"r1_mean": float(np.mean(r1_total)) if r1_total else 0,
            "ctrl_mean": float(np.mean(ctrl_total)) if ctrl_total else 0}


# ═══════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════
def main():
    print("=" * 60)
    print("MUNICIPALITY-LEVEL CRIME ANALYSIS")
    print("=" * 60)

    with open(os.path.join(DATA_DIR, "round1_japan_stores.json")) as f:
        stores = json.load(f)

    # Tokyo analysis
    print("\n--- Tokyo ---")
    all_data = load_all_tokyo_crime()
    print(f"Loaded {len(all_data)} years of Tokyo municipality data")
    for y in sorted(all_data.keys()):
        n_muni = len(all_data[y])
        total = sum(v["total"] for v in all_data[y].values())
        print(f"  {y}: {n_muni} municipalities, {total:,} total crimes")

    df, r1_munis = build_tokyo_panel(all_data, stores)
    print(f"\nPanel: {len(df)} obs, {df['municipality'].nunique()} municipalities, "
          f"{df['year'].nunique()} years")
    print(f"R1 municipalities: {r1_munis}")

    did_results = run_tokyo_did(df)
    plot_tokyo_trends(df, r1_munis)
    plot_tokyo_forest(did_results)
    plot_individual_municipalities(df, r1_munis)

    # Neighbor municipality analysis
    neighbor_results = run_neighbor_municipality_did(df, r1_munis)
    plot_neighbor_municipality_trends(df, r1_munis)

    # Osaka snapshot
    print("\n--- Osaka ---")
    osaka_result = analyze_osaka_snapshot(stores)

    # Save panel data
    df.to_csv(os.path.join(OUTPUT_DIR, "tokyo_municipality_panel.csv"),
              index=False, encoding="utf-8")

    print("\n" + "=" * 60)
    print("MUNICIPALITY ANALYSIS COMPLETE")
    print("=" * 60)


if __name__ == "__main__":
    main()
