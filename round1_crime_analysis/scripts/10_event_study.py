"""
10_event_study.py
Event Study + ITS + DiD for Round1 crime analysis (Japan & US).

Analyses:
  1. US Event Study: Staggered adoption design, event-time dummies (t-3..t+3)
  2. US ITS: Level + slope change at R1 opening for treated counties
  3. US Classical DiD: Pre/post × R1/control
  4. Japan (Tokyo) Event Study: Municipality-level, 2009-2025
  5. Japan (Tokyo) ITS: Pre/post opening
  6. Japan (Nationwide) ITS: R1 cities theft trend vs controls (2018-2023)
"""

import os
import sys
import json
import csv
import warnings
import time

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import statsmodels.formula.api as smf
import statsmodels.api as sm

warnings.filterwarnings("ignore", category=FutureWarning)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
FIG_DIR = os.path.join(BASE_DIR, "figures")
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(FIG_DIR, exist_ok=True)

C_R1 = "#E63946"
C_CTRL = "#457B9D"
C_ACCENT = "#2A9D8F"

# ═══════════════════════════════════════════════════════════════════
# US DATA
# ═══════════════════════════════════════════════════════════════════

def load_us_panel():
    """Load US county panel from 09_us_county_analysis output."""
    path = os.path.join(OUTPUT_DIR, "us_county_panel.csv")
    if not os.path.exists(path):
        print("ERROR: Run 09_us_county_analysis.py first.")
        return None
    df = pd.read_csv(path)
    return df


def us_event_study(panel):
    """Staggered event study: event-time dummies relative to first R1 opening."""
    print(f"\n{'='*60}")
    print("US EVENT STUDY (Staggered Adoption)")
    print(f"{'='*60}")

    # Only counties that opened R1 within our data window (need pre+post)
    treated = panel[(panel["has_r1"] == 1) & (panel["first_r1_year"] >= 2015) & (panel["first_r1_year"] <= 2018)].copy()
    control = panel[panel["has_r1"] == 0].copy()

    if len(treated) == 0:
        print("No treated counties with pre+post data.")
        return {}

    # Create event time for treated
    treated["event_time"] = treated["year"] - treated["first_r1_year"]
    # Keep event_time in [-3, +3] range
    treated = treated[(treated["event_time"] >= -3) & (treated["event_time"] <= 3)].copy()

    # For controls, assign pseudo event time using average opening year
    avg_open = int(treated["first_r1_year"].mean())
    control["first_r1_year"] = avg_open
    control["event_time"] = control["year"] - avg_open
    control = control[(control["event_time"] >= -3) & (control["event_time"] <= 3)].copy()

    combined = pd.concat([treated, control], ignore_index=True)
    combined["treated"] = combined["has_r1"]

    print(f"Treated counties: {treated['county_key'].nunique()} (opened 2015-2018)")
    print(f"Control counties: {control['county_key'].nunique()}")
    print(f"Event time range: [{combined['event_time'].min()}, {combined['event_time'].max()}]")

    results = {}
    crime_categories = [
        ("Property Crime", "property_crime"),
        ("Burglary", "burglary"),
        ("Larceny-Theft", "larceny_theft"),
        ("Violent Crime", "violent_crime"),
    ]

    for title, col in crime_categories:
        if col not in combined.columns:
            continue

        reg = combined[["county_key", "year", "event_time", "treated", col]].dropna().copy()
        reg = reg.rename(columns={col: "y"})

        # Create event-time dummies (omit t=-1 as reference)
        # Use 'm' prefix for negative event times to avoid patsy formula issues
        def et_name(t):
            return f"et_m{abs(t)}" if t < 0 else f"et_p{t}"

        for t in range(-3, 4):
            if t == -1:
                continue
            reg[et_name(t)] = ((reg["event_time"] == t) & (reg["treated"] == 1)).astype(int)

        et_vars = [et_name(t) for t in range(-3, 4) if t != -1]
        formula = f"y ~ {' + '.join(et_vars)} + C(county_key) + C(year)"

        try:
            m = smf.ols(formula, data=reg).fit(
                cov_type="cluster", cov_kwds={"groups": reg["county_key"]}
            )
            coefs = {}
            for t in range(-3, 4):
                if t == -1:
                    coefs[t] = {"coef": 0, "se": 0, "p": 1.0, "ci": [0, 0]}
                else:
                    var = et_name(t)
                    if var in m.params:
                        ci = m.conf_int().loc[var].tolist()
                        coefs[t] = {
                            "coef": float(m.params[var]),
                            "se": float(m.bse[var]),
                            "p": float(m.pvalues[var]),
                            "ci": [float(ci[0]), float(ci[1])],
                        }

            results[title] = coefs
            post_coefs = [coefs[t]["coef"] for t in range(0, 4) if t in coefs]
            pre_coefs = [coefs[t]["coef"] for t in range(-3, -1) if t in coefs]
            print(f"\n  {title}:")
            print(f"    Pre-trend avg: {np.mean(pre_coefs):.1f}")
            print(f"    Post-trend avg: {np.mean(post_coefs):.1f}")
        except Exception as e:
            print(f"  {title}: failed — {e}")

    return results


def us_its(panel):
    """ITS analysis for US R1 counties: level + slope change at opening."""
    print(f"\n{'='*60}")
    print("US ITS (Interrupted Time Series)")
    print(f"{'='*60}")

    # Use counties with pre+post data
    treated = panel[(panel["has_r1"] == 1) & (panel["first_r1_year"] >= 2015) & (panel["first_r1_year"] <= 2018)].copy()
    treated["event_time"] = treated["year"] - treated["first_r1_year"]
    treated["post"] = (treated["event_time"] >= 0).astype(int)
    treated["time_since"] = treated["event_time"].clip(lower=0)

    results = {}
    for title, col in [
        ("Property Crime", "property_crime"),
        ("Burglary", "burglary"),
        ("Larceny-Theft", "larceny_theft"),
        ("Violent Crime", "violent_crime"),
    ]:
        if col not in treated.columns:
            continue

        reg = treated[["county_key", "year", "event_time", "post", "time_since", col]].dropna().copy()
        reg = reg.rename(columns={col: "y"})

        try:
            m = smf.ols(
                "y ~ event_time + post + time_since + C(county_key)", data=reg
            ).fit(cov_type="cluster", cov_kwds={"groups": reg["county_key"]})

            level_change = float(m.params.get("post", 0))
            level_p = float(m.pvalues.get("post", 1))
            slope_change = float(m.params.get("time_since", 0))
            slope_p = float(m.pvalues.get("time_since", 1))

            sig_l = "*" if level_p < 0.05 else ""
            sig_s = "*" if slope_p < 0.05 else ""
            print(f"\n  {title}:")
            print(f"    Level change at opening: {level_change:+.1f} {sig_l} (p={level_p:.4f})")
            print(f"    Slope change post-opening: {slope_change:+.1f} {sig_s} (p={slope_p:.4f})")

            results[title] = {
                "level_change": round(level_change, 2),
                "level_p": round(level_p, 4),
                "slope_change": round(slope_change, 2),
                "slope_p": round(slope_p, 4),
                "n_counties": int(reg["county_key"].nunique()),
                "n_obs": int(len(reg)),
            }
        except Exception as e:
            print(f"  {title}: failed — {e}")

    return results


def us_classical_did(panel):
    """Classical pre/post DiD for US counties."""
    print(f"\n{'='*60}")
    print("US CLASSICAL DiD (Pre/Post × R1/Control)")
    print(f"{'='*60}")

    treated = panel[(panel["has_r1"] == 1) & (panel["first_r1_year"] >= 2015) & (panel["first_r1_year"] <= 2018)].copy()
    control = panel[panel["has_r1"] == 0].copy()

    treated["event_time"] = treated["year"] - treated["first_r1_year"]
    treated["post"] = (treated["event_time"] >= 0).astype(int)

    avg_open = int(treated["first_r1_year"].mean())
    control["post"] = (control["year"] >= avg_open).astype(int)

    combined = pd.concat([treated, control], ignore_index=True)
    combined["treated"] = combined["has_r1"]
    combined["did"] = combined["treated"] * combined["post"]

    results = {}
    for title, col in [
        ("Property Crime", "property_crime"),
        ("Burglary", "burglary"),
        ("Larceny-Theft", "larceny_theft"),
        ("Violent Crime", "violent_crime"),
        ("Motor Vehicle Theft", "motor_vehicle_theft"),
        ("Robbery", "robbery"),
        ("Murder", "murder"),
        ("Aggravated Assault", "aggravated_assault"),
    ]:
        if col not in combined.columns:
            continue

        reg = combined[["county_key", "year", "treated", "post", "did", col]].dropna().copy()
        reg = reg.rename(columns={col: "y"})

        try:
            m = smf.ols("y ~ did + C(county_key) + C(year)", data=reg).fit(
                cov_type="cluster", cov_kwds={"groups": reg["county_key"]}
            )
            coef = float(m.params["did"])
            p = float(m.pvalues["did"])
            ci = m.conf_int().loc["did"].tolist()
            sig = "***" if p < 0.001 else "**" if p < 0.01 else "*" if p < 0.05 else ""
            print(f"  {title}: DiD = {coef:+.1f} {sig} (p={p:.4f})  [{ci[0]:.1f}, {ci[1]:.1f}]")

            results[title] = {
                "did_coefficient": round(coef, 2),
                "p_value": round(p, 4),
                "ci_95": [round(ci[0], 2), round(ci[1], 2)],
                "n_treated": int(reg[reg["treated"] == 1]["county_key"].nunique()),
                "n_control": int(reg[reg["treated"] == 0]["county_key"].nunique()),
            }
        except Exception as e:
            print(f"  {title}: failed — {e}")

    return results


# ═══════════════════════════════════════════════════════════════════
# JAPAN (TOKYO) DATA
# ═══════════════════════════════════════════════════════════════════

CSV_YEARS = {
    "H29": 2017, "H30": 2018,
    "R2": 2020, "R3": 2021, "R4": 2022, "R5": 2023, "R6": 2024, "R7": 2025,
}


def parse_tokyo_csv(path, year):
    municipality_totals = {}
    with open(path, "r", encoding="cp932") as f:
        reader = csv.reader(f)
        header = next(reader)
        for row in reader:
            loc = row[0].strip()
            if "計" in loc or "不詳" in loc or "他県" in loc or "海外" in loc:
                continue
            try:
                total = int(row[1])
            except (ValueError, IndexError):
                continue
            violent = int(row[2]) if len(row) > 2 and row[2].strip() else 0
            rough = int(row[5]) if len(row) > 5 and row[5].strip() else 0
            burglary = int(row[11]) if len(row) > 11 and row[11].strip() else 0
            theft_non = int(row[20]) if len(row) > 20 and row[20].strip() else 0
            other = int(row[32]) if len(row) > 32 and row[32].strip() else 0

            municipality = None
            for suffix in ["区", "市", "町", "村"]:
                idx = loc.find(suffix)
                if idx >= 0:
                    municipality = loc[:idx + 1]
                    break

            if municipality:
                if municipality not in municipality_totals:
                    municipality_totals[municipality] = {
                        "total": 0, "violent": 0, "rough": 0,
                        "burglary": 0, "theft_non": 0, "other": 0,
                    }
                m = municipality_totals[municipality]
                m["total"] += total
                m["violent"] += violent
                m["rough"] += rough
                m["burglary"] += burglary
                m["theft_non"] += theft_non
                m["other"] += other
    return municipality_totals


def parse_tokyo_xls(path, year):
    import xlrd
    wb = xlrd.open_workbook(path)
    municipality_totals = {}
    skip_sheets = {"最終", "合計", "総計"}
    for sname in wb.sheet_names():
        if sname in skip_sheets:
            continue
        ws = wb.sheet_by_name(sname)
        totals = {"total": 0, "violent": 0, "rough": 0,
                  "burglary": 0, "theft_non": 0, "other": 0}
        for r in range(7, ws.nrows):
            loc = str(ws.cell_value(r, 0)).strip()
            if "計" in loc or "不詳" in loc or not loc:
                continue
            def cell_int(c):
                v = ws.cell_value(r, c)
                return int(v) if isinstance(v, (int, float)) and v > 0 else 0
            totals["total"] += cell_int(1)
            totals["violent"] += cell_int(2)
            totals["rough"] += cell_int(5)
            totals["burglary"] += cell_int(11)
            totals["theft_non"] += cell_int(20)
            totals["other"] += cell_int(32)
        municipality_totals[sname] = totals
    return municipality_totals


def parse_tokyo_xlsx(path, year):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    municipality_totals = {}
    skip_sheets = {"最終", "合計", "総計"}
    for sname in wb.sheetnames:
        if sname in skip_sheets:
            continue
        ws = wb[sname]
        totals = {"total": 0, "violent": 0, "rough": 0,
                  "burglary": 0, "theft_non": 0, "other": 0}
        for row in ws.iter_rows(min_row=8, values_only=True):
            loc = str(row[0]).strip() if row[0] else ""
            if "計" in loc or "不詳" in loc or not loc:
                continue
            def cell_int(idx):
                v = row[idx] if idx < len(row) else 0
                return int(v) if isinstance(v, (int, float)) and v and v > 0 else 0
            totals["total"] += cell_int(1)
            totals["violent"] += cell_int(2)
            totals["rough"] += cell_int(5)
            totals["burglary"] += cell_int(11)
            totals["theft_non"] += cell_int(20)
            totals["other"] += cell_int(32)
        municipality_totals[sname] = totals
    return municipality_totals


def load_all_tokyo_crime():
    tokyo_dir = os.path.join(DATA_DIR, "tokyo_crime")
    all_data = {}
    xls_years = {"H21": 2009, "H22": 2010, "H23": 2011, "H24": 2012, "H25": 2013}
    for fname, year in xls_years.items():
        path = os.path.join(tokyo_dir, f"{fname}.xls")
        if os.path.exists(path):
            all_data[year] = parse_tokyo_xls(path, year)
    # Only H28.xlsx is a valid xlsx; H26/H27/R1 .xlsx are HTML files
    xlsx_path = os.path.join(tokyo_dir, "H28.xlsx")
    if os.path.exists(xlsx_path):
        all_data[2016] = parse_tokyo_xlsx(xlsx_path, 2016)
    # CSV files: require >50KB to skip HTML redirect pages (~11KB)
    for fname, year in CSV_YEARS.items():
        path = os.path.join(tokyo_dir, f"{fname}.csv")
        if os.path.exists(path) and os.path.getsize(path) > 50000:
            all_data[year] = parse_tokyo_csv(path, year)
    return all_data


def build_tokyo_panel(all_data):
    with open(os.path.join(DATA_DIR, "round1_japan_stores.json")) as f:
        stores = json.load(f)

    tokyo_stores = [s for s in stores if s.get("prefecture") == "東京都"]
    r1_munis = {}
    for s in tokyo_stores:
        city = s.get("city", "")
        open_year = s.get("open_year", 9999)
        for suffix in ["区", "市", "町", "村"]:
            idx = city.find(suffix)
            if idx >= 0:
                muni = city[:idx + 1]
                if muni not in r1_munis or open_year < r1_munis[muni]:
                    r1_munis[muni] = open_year
                break

    rows = []
    for year, muni_data in sorted(all_data.items()):
        for muni, crimes in muni_data.items():
            first_r1 = r1_munis.get(muni, np.nan)
            rows.append({
                "year": year,
                "municipality": muni,
                "total_crimes": crimes["total"],
                "violent": crimes["violent"],
                "rough": crimes["rough"],
                "burglary": crimes["burglary"],
                "theft_non": crimes["theft_non"],
                "other": crimes["other"],
                "has_r1": 1 if muni in r1_munis else 0,
                "first_r1_year": first_r1,
            })

    df = pd.DataFrame(rows)
    df["event_time"] = df["year"] - df["first_r1_year"]
    df["post"] = (df["has_r1"] == 1) & (df["year"] >= df["first_r1_year"])
    df["post"] = df["post"].astype(int)
    return df, r1_munis


def jp_tokyo_event_study(panel):
    """Event study for Tokyo municipalities."""
    print(f"\n{'='*60}")
    print("JAPAN (TOKYO) EVENT STUDY")
    print(f"{'='*60}")

    treated = panel[(panel["has_r1"] == 1)].copy()
    control = panel[panel["has_r1"] == 0].copy()

    treated = treated[(treated["event_time"] >= -5) & (treated["event_time"] <= 10)].copy()
    avg_open = int(treated["first_r1_year"].mean())
    control["first_r1_year"] = avg_open
    control["event_time"] = control["year"] - avg_open
    control = control[(control["event_time"] >= -5) & (control["event_time"] <= 10)].copy()

    combined = pd.concat([treated, control], ignore_index=True)
    combined["treated"] = combined["has_r1"]

    print(f"Treated municipalities: {treated['municipality'].nunique()}")
    print(f"Control municipalities: {control['municipality'].nunique()}")
    print(f"Years: {sorted(panel['year'].unique())}")

    # Encode municipality as numeric ID (patsy has trouble with CJK in C())
    muni_map = {m: i for i, m in enumerate(combined["municipality"].unique())}
    combined["muni_id"] = combined["municipality"].map(muni_map)

    results = {}
    for title, col in [
        ("Total Crimes", "total_crimes"),
        ("Violent Crime", "violent"),
        ("Burglary", "burglary"),
        ("Non-Intrusion Theft", "theft_non"),
    ]:
        if col not in combined.columns:
            continue
        reg = combined[["municipality", "muni_id", "year", "event_time", "treated", col]].dropna().copy()
        reg = reg.rename(columns={col: "y"})

        et_range = sorted(reg["event_time"].unique())
        ref = -1
        et_vars = []
        for t in et_range:
            if t == ref:
                continue
            ti = int(t)
            vname = f"et_m{abs(ti)}" if ti < 0 else f"et_p{ti}"
            reg[vname] = ((reg["event_time"] == t) & (reg["treated"] == 1)).astype(int)
            et_vars.append(vname)

        formula = f"y ~ {' + '.join(et_vars)} + C(muni_id) + C(year)"
        try:
            m = smf.ols(formula, data=reg).fit(
                cov_type="cluster", cov_kwds={"groups": reg["municipality"]}
            )
            coefs = {}
            for t in et_range:
                if t == ref:
                    coefs[t] = {"coef": 0, "se": 0, "p": 1.0, "ci": [0, 0]}
                else:
                    ti = int(t)
                    vname = f"et_m{abs(ti)}" if ti < 0 else f"et_p{ti}"
                    if vname in m.params:
                        ci = m.conf_int().loc[vname].tolist()
                        coefs[t] = {
                            "coef": float(m.params[vname]),
                            "se": float(m.bse[vname]),
                            "p": float(m.pvalues[vname]),
                            "ci": [float(ci[0]), float(ci[1])],
                        }
            results[title] = coefs
            pre = [coefs[t]["coef"] for t in et_range if t < -1 and t in coefs]
            post = [coefs[t]["coef"] for t in et_range if t >= 0 and t in coefs]
            print(f"\n  {title}:")
            print(f"    Pre-trend avg: {np.mean(pre):.1f}" if pre else "    No pre-trend data")
            print(f"    Post-trend avg: {np.mean(post):.1f}" if post else "    No post-trend data")
        except Exception as e:
            print(f"  {title}: failed — {e}")

    return results


def jp_tokyo_its(panel):
    """ITS for Tokyo R1 municipalities."""
    print(f"\n{'='*60}")
    print("JAPAN (TOKYO) ITS")
    print(f"{'='*60}")

    treated = panel[panel["has_r1"] == 1].copy()
    treated["event_time"] = treated["year"] - treated["first_r1_year"]
    treated["time_since"] = treated["event_time"].clip(lower=0)

    muni_map = {m: i for i, m in enumerate(treated["municipality"].unique())}
    treated["muni_id"] = treated["municipality"].map(muni_map)

    results = {}
    for title, col in [
        ("Total Crimes", "total_crimes"),
        ("Violent Crime", "violent"),
        ("Burglary", "burglary"),
        ("Non-Intrusion Theft", "theft_non"),
    ]:
        if col not in treated.columns:
            continue
        reg = treated[["municipality", "muni_id", "year", "event_time", "post", "time_since", col]].dropna().copy()
        reg = reg.rename(columns={col: "y"})

        try:
            m = smf.ols("y ~ event_time + post + time_since + C(muni_id)", data=reg).fit(
                cov_type="cluster", cov_kwds={"groups": reg["municipality"]}
            )
            level = float(m.params.get("post", 0))
            level_p = float(m.pvalues.get("post", 1))
            slope = float(m.params.get("time_since", 0))
            slope_p = float(m.pvalues.get("time_since", 1))
            sig_l = "*" if level_p < 0.05 else ""
            sig_s = "*" if slope_p < 0.05 else ""
            print(f"\n  {title}:")
            print(f"    Level change: {level:+.1f} {sig_l} (p={level_p:.4f})")
            print(f"    Slope change: {slope:+.1f} {sig_s} (p={slope_p:.4f})")

            results[title] = {
                "level_change": round(level, 2),
                "level_p": round(level_p, 4),
                "slope_change": round(slope, 2),
                "slope_p": round(slope_p, 4),
                "n_municipalities": int(reg["municipality"].nunique()),
            }
        except Exception as e:
            print(f"  {title}: failed — {e}")

    return results


def jp_tokyo_did(panel):
    """Classical DiD for Tokyo municipalities."""
    print(f"\n{'='*60}")
    print("JAPAN (TOKYO) CLASSICAL DiD")
    print(f"{'='*60}")

    treated = panel[panel["has_r1"] == 1].copy()
    control = panel[panel["has_r1"] == 0].copy()
    avg_open = int(treated["first_r1_year"].mean())
    control["post"] = (control["year"] >= avg_open).astype(int)

    combined = pd.concat([treated, control], ignore_index=True)
    combined["treated"] = combined["has_r1"]
    combined["did"] = combined["treated"] * combined["post"]

    muni_map = {m: i for i, m in enumerate(combined["municipality"].unique())}
    combined["muni_id"] = combined["municipality"].map(muni_map)

    results = {}
    for title, col in [
        ("Total Crimes", "total_crimes"),
        ("Violent Crime", "violent"),
        ("Burglary", "burglary"),
        ("Non-Intrusion Theft", "theft_non"),
    ]:
        if col not in combined.columns:
            continue
        reg = combined[["municipality", "muni_id", "year", "treated", "post", "did", col]].dropna().copy()
        reg = reg.rename(columns={col: "y"})

        try:
            m = smf.ols("y ~ did + C(muni_id) + C(year)", data=reg).fit(
                cov_type="cluster", cov_kwds={"groups": reg["municipality"]}
            )
            coef = float(m.params["did"])
            p = float(m.pvalues["did"])
            ci = m.conf_int().loc["did"].tolist()
            sig = "***" if p < 0.001 else "**" if p < 0.01 else "*" if p < 0.05 else ""
            print(f"  {title}: DiD = {coef:+.1f} {sig} (p={p:.4f})  [{ci[0]:.1f}, {ci[1]:.1f}]")
            results[title] = {
                "did_coefficient": round(coef, 2),
                "p_value": round(p, 4),
                "ci_95": [round(ci[0], 2), round(ci[1], 2)],
            }
        except Exception as e:
            print(f"  {title}: failed — {e}")

    return results


# ═══════════════════════════════════════════════════════════════════
# PLOTTING
# ═══════════════════════════════════════════════════════════════════

def plot_event_study(results, country, filename):
    """Event study coefficient plot."""
    n_panels = len(results)
    if n_panels == 0:
        return

    fig, axes = plt.subplots(1, n_panels, figsize=(5 * n_panels, 5))
    if n_panels == 1:
        axes = [axes]

    for idx, (title, coefs) in enumerate(results.items()):
        ax = axes[idx]
        times = sorted(coefs.keys())
        vals = [coefs[t]["coef"] for t in times]
        ci_lo = [coefs[t]["ci"][0] for t in times]
        ci_hi = [coefs[t]["ci"][1] for t in times]

        ax.fill_between(times, ci_lo, ci_hi, alpha=0.2, color=C_R1)
        ax.plot(times, vals, color=C_R1, marker="o", markersize=5, lw=2)
        ax.axhline(0, color="black", ls="--", alpha=0.5)
        ax.axvline(-0.5, color="gray", ls=":", alpha=0.5)

        ax.set_xlabel("Event Time (years relative to R1 opening)", fontsize=9)
        ax.set_ylabel("Coefficient (vs t=-1)", fontsize=9)
        ax.set_title(title, fontsize=10, fontweight="bold")

        # Annotate pre/post
        ax.text(min(times) + 0.3, ax.get_ylim()[1] * 0.9, "Pre", fontsize=8, color="gray")
        ax.text(0.3, ax.get_ylim()[1] * 0.9, "Post", fontsize=8, color=C_R1)

    country_label = "US Counties" if country == "US" else "Tokyo Municipalities"
    fig.suptitle(f"Event Study: R1 Opening Effect on Crime ({country_label})",
                 fontsize=12, fontweight="bold")
    plt.tight_layout(rect=[0, 0, 1, 0.93])
    path = os.path.join(FIG_DIR, filename)
    plt.savefig(path, dpi=200, bbox_inches="tight")
    plt.close()
    print(f"[{filename}] Event study saved.")


def plot_its_trends(panel, country, r1_col, id_col, crime_col, title, filename):
    """ITS-style time series with pre/post shading."""
    treated = panel[panel[r1_col] == 1].copy()
    control = panel[panel[r1_col] == 0].copy()

    if "first_r1_year" in treated.columns:
        treated["event_time"] = treated["year"] - treated["first_r1_year"]
    else:
        treated["event_time"] = treated["year"]

    fig, axes = plt.subplots(1, 2, figsize=(14, 5))

    # Left: Raw time series
    ax = axes[0]
    for grp, df, color, label in [
        (1, treated, C_R1, "R1"),
        (0, control, C_CTRL, "Control"),
    ]:
        g = df.groupby("year")[crime_col].mean()
        ax.plot(g.index, g.values, color=color, marker="o", markersize=4, lw=2, label=label)

    if "first_r1_year" in treated.columns:
        avg_open = treated["first_r1_year"].mean()
        ax.axvline(avg_open, color="red", ls="--", alpha=0.6, label=f"Avg R1 opening ({avg_open:.0f})")

    ax.set_xlabel("Year", fontsize=10)
    ax.set_ylabel(f"Mean {title}", fontsize=10)
    ax.set_title(f"{title}: R1 vs Control", fontsize=11, fontweight="bold")
    ax.legend(fontsize=8)

    # Right: Event-time view (treated only)
    ax = axes[1]
    if "event_time" in treated.columns and treated["event_time"].notna().any():
        g = treated.groupby("event_time")[crime_col].mean()
        pre = g[g.index < 0]
        post = g[g.index >= 0]

        if len(pre) > 0:
            ax.plot(pre.index, pre.values, color=C_CTRL, marker="o", markersize=4, lw=2, label="Pre-opening")
        if len(post) > 0:
            ax.plot(post.index, post.values, color=C_R1, marker="o", markersize=4, lw=2, label="Post-opening")
        ax.axvline(-0.5, color="red", ls="--", alpha=0.6)

        # Fit pre/post trend lines
        if len(pre) >= 2:
            z = np.polyfit(pre.index, pre.values, 1)
            ax.plot(pre.index, np.polyval(z, pre.index), color=C_CTRL, ls="--", alpha=0.5)
        if len(post) >= 2:
            z = np.polyfit(post.index, post.values, 1)
            ax.plot(post.index, np.polyval(z, post.index), color=C_R1, ls="--", alpha=0.5)

    ax.set_xlabel("Event Time (years since R1 opening)", fontsize=10)
    ax.set_ylabel(f"Mean {title}", fontsize=10)
    ax.set_title(f"{title}: Pre/Post Opening (R1 Only)", fontsize=11, fontweight="bold")
    ax.legend(fontsize=8)

    fig.suptitle(f"ITS Analysis: {country}", fontsize=13, fontweight="bold")
    plt.tight_layout(rect=[0, 0, 1, 0.93])
    path = os.path.join(FIG_DIR, filename)
    plt.savefig(path, dpi=200, bbox_inches="tight")
    plt.close()
    print(f"[{filename}] ITS trends saved.")


def plot_did_comparison(us_did, jp_did, filename):
    """Side-by-side DiD coefficients for US and Japan."""
    fig, axes = plt.subplots(1, 2, figsize=(14, 6))

    for ax, results, country in [(axes[0], us_did, "US Counties"), (axes[1], jp_did, "Tokyo Municipalities")]:
        if not results:
            ax.text(0.5, 0.5, "No data", ha="center", va="center", fontsize=12)
            ax.set_title(f"DiD: {country}", fontweight="bold")
            continue

        cats = list(results.keys())
        coefs = [results[c]["did_coefficient"] for c in cats]
        cis = [results[c]["ci_95"] for c in cats]
        ps = [results[c]["p_value"] for c in cats]

        for i, (cat, coef, ci, p) in enumerate(zip(cats, coefs, cis, ps)):
            color = C_R1 if p < 0.05 else "#999999"
            ax.errorbar(coef, i, xerr=[[coef - ci[0]], [ci[1] - coef]],
                        fmt="o", color=color, markersize=8, capsize=5, capthick=2, lw=2)
            sig = " ***" if p < 0.001 else " **" if p < 0.01 else " *" if p < 0.05 else ""
            ax.annotate(f"p={p:.3f}{sig}", xy=(ci[1] + abs(ci[1] - ci[0]) * 0.1, i),
                        fontsize=7, va="center")

        ax.axvline(0, color="black", ls="--", alpha=0.5)
        ax.set_yticks(range(len(cats)))
        ax.set_yticklabels(cats, fontsize=9)
        ax.set_xlabel("DiD Coefficient", fontsize=10)
        ax.set_title(f"DiD: {country}", fontsize=11, fontweight="bold")
        ax.invert_yaxis()

    plt.suptitle("Classical DiD: Round1 Effect on Crime\nUS Counties (2014–2019) vs Tokyo Municipalities (2009–2025)",
                 fontsize=12, fontweight="bold")
    plt.tight_layout(rect=[0, 0, 1, 0.91])
    path = os.path.join(FIG_DIR, filename)
    plt.savefig(path, dpi=200, bbox_inches="tight")
    plt.close()
    print(f"[{filename}] DiD comparison saved.")


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════

def main():
    print("=" * 60)
    print("EVENT STUDY + ITS + DiD ANALYSIS")
    print("=" * 60)

    all_results = {}

    # --- US ---
    us_panel = load_us_panel()
    if us_panel is not None:
        us_es = us_event_study(us_panel)
        us_its_res = us_its(us_panel)
        us_did = us_classical_did(us_panel)

        all_results["us_event_study"] = us_es
        all_results["us_its"] = us_its_res
        all_results["us_did"] = us_did

        if us_es:
            plot_event_study(us_es, "US", "fig_us04_event_study.png")

        # ITS plots for key categories
        treated_counties = us_panel[
            (us_panel["has_r1"] == 1) &
            (us_panel["first_r1_year"] >= 2015) &
            (us_panel["first_r1_year"] <= 2018)
        ].copy()
        treated_counties["event_time"] = treated_counties["year"] - treated_counties["first_r1_year"]
        ctrl_counties = us_panel[us_panel["has_r1"] == 0].copy()
        avg_open = int(treated_counties["first_r1_year"].mean())
        ctrl_counties["first_r1_year"] = avg_open
        ctrl_counties["event_time"] = ctrl_counties["year"] - avg_open
        us_its_panel = pd.concat([treated_counties, ctrl_counties], ignore_index=True)
        plot_its_trends(us_its_panel, "US Counties", "has_r1", "county_key",
                        "property_crime", "Property Crime", "fig_us05_its_property.png")

    # --- Japan (Tokyo) ---
    print("\n\nLoading Tokyo municipality data...")
    all_data = load_all_tokyo_crime()
    if all_data:
        jp_panel, r1_munis = build_tokyo_panel(all_data)
        print(f"Tokyo panel: {len(jp_panel)} obs, {jp_panel['municipality'].nunique()} municipalities, "
              f"{jp_panel['year'].nunique()} years")
        print(f"R1 municipalities: {jp_panel[jp_panel['has_r1']==1]['municipality'].unique().tolist()}")

        jp_es = jp_tokyo_event_study(jp_panel)
        jp_its_res = jp_tokyo_its(jp_panel)
        jp_did = jp_tokyo_did(jp_panel)

        all_results["jp_event_study"] = jp_es
        all_results["jp_its"] = jp_its_res
        all_results["jp_did"] = jp_did

        if jp_es:
            plot_event_study(jp_es, "JP", "fig_jp19_event_study.png")

        plot_its_trends(jp_panel, "Tokyo Municipalities", "has_r1", "municipality",
                        "total_crimes", "Total Crimes", "fig_jp20_its_total.png")
    else:
        jp_did = {}

    # --- Combined DiD comparison ---
    plot_did_comparison(
        all_results.get("us_did", {}),
        all_results.get("jp_did", {}),
        "fig_combined_did_comparison.png",
    )

    # Save all results
    # Convert numpy types for JSON serialization
    def convert(obj):
        if isinstance(obj, (np.integer,)):
            return int(obj)
        if isinstance(obj, (np.floating,)):
            return float(obj)
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        return obj

    class NpEncoder(json.JSONEncoder):
        def default(self, obj):
            return convert(obj)

    results_path = os.path.join(OUTPUT_DIR, "event_study_results.json")
    with open(results_path, "w") as f:
        json.dump(all_results, f, indent=2, cls=NpEncoder)

    print(f"\n{'='*60}")
    print("EVENT STUDY + ITS + DiD ANALYSIS COMPLETE")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
