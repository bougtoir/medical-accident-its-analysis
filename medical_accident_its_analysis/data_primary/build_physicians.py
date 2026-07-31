#!/usr/bin/env python3
"""Build biennial physician counts by specialty from official MHLW
'医師・歯科医師・薬剤師統計' files, aggregated to the Supreme Court
litigation specialty categories.

Primary sources (raw files under data_primary/raw/physicians/):
  - 2018 統計表 file, sheet '参考２':
      (参考2) 医療施設従事医師数の年次推移，主たる診療科別
      -> gives ALL specialties for 2000,2002,...,2018 in one harmonised table.
  - 2020/2022/2024 概況 files, sheet '表４':
      表4 主たる診療科別にみた医療施設に従事する医師数
      -> single-year all-specialty counts for the years not in 参考2.

MHLW note carried from 参考2: '診療科別医師数の年次推移については、標榜診療科名の
改正の影響等により、単純な比較が難しい場合がある' (2008 標榜科 reform). To obtain a
CONSISTENT broad series we sum every specialty row belonging to a Court
category for each year (pre-2008 the subspecialty rows are '・'/absent and the
broad parent row already carries them, so the per-year sum is comparable).

Mapping decisions (documented; ambiguous ones flagged for sensitivity):
  内科(broad) = 内科 + all internal-medicine subspecialties (腎臓/糖尿病/血液/
                呼吸器/循環器/消化器/神経内科/脳神経内科/アレルギー/リウマチ/
                感染症/心療内科). Excludes 皮膚科, 精神科 (separate Court cats).
  外科(broad) = 外科 + 乳腺/消化器/肛門(こう門)/気管食道/呼吸器/心臓血管/小児外科.
                Excludes 脳神経外科 (no Court category; likely その他/外科) and
                整形/形成/美容外科 (separate cats). -> flagged for sensitivity.
  形成外科     = 形成外科 + 美容外科 (cosmetic litigation reported together).
  精神科       = 精神科 + 神経科(pre-2008 label).
  産婦人科     = 産婦人科 + 産科 + 婦人科.
  others map 1:1 (整形外科, 小児科, 眼科, 耳鼻咽喉科, 泌尿器科, 皮膚科, 麻酔科).
"""
import os, re, json, hashlib
import xlrd
import openpyxl
import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
RAW = os.path.join(HERE, "raw", "physicians")

# base(label): text before first paren, trailing footnote digits/)/space removed
def base(label):
    if not isinstance(label, str):
        return ""
    s = re.split(r"[（(]", label)[0]
    s = s.replace("\u3000", "").replace(" ", "")
    s = re.sub(r"[0-9)\)]+$", "", s)
    return s.strip()

INTERNAL = {"内科", "腎臓内科", "糖尿病内科", "血液内科", "呼吸器科", "呼吸器内科",
            "循環器科", "循環器内科", "消化器科", "消化器内科", "神経内科",
            "脳神経内科", "アレルギー科", "リウマチ科", "感染症内科", "心療内科"}
PSYCH = {"精神科", "神経科"}
SURG = {"外科", "乳腺外科", "消化器外科", "こう門科", "肛門外科", "気管食道科",
        "気管食道外科", "呼吸器外科", "心臓血管外科", "小児外科"}
PLASTIC = {"形成外科", "美容外科"}
OBGYN = {"産婦人科", "産科", "婦人科"}

MAP = {}
for b in INTERNAL: MAP[b] = "内科"
for b in PSYCH: MAP[b] = "精神科"
for b in SURG: MAP[b] = "外科"
for b in PLASTIC: MAP[b] = "形成外科"
for b in OBGYN: MAP[b] = "産婦人科"
MAP["整形外科"] = "整形外科"
MAP["小児科"] = "小児科"
MAP["眼科"] = "眼科"
MAP["耳鼻いんこう科"] = "耳鼻咽喉科"
MAP["耳鼻咽喉科"] = "耳鼻咽喉科"
MAP["泌尿器科"] = "泌尿器科"
MAP["皮膚科"] = "皮膚科"
MAP["麻酔科"] = "麻酔科"

CORE = ["内科", "外科", "整形外科", "形成外科", "産婦人科", "小児科", "精神科",
        "眼科", "耳鼻咽喉科", "泌尿器科", "皮膚科", "麻酔科"]


def _num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.replace(" ", "").replace("\u3000", "").replace(",", "")
        if re.fullmatch(r"[0-9]+(\.[0-9]+)?", s):
            return float(s)
    return None


def parse_sanko2(path):
    """(参考2) year-series table -> {year: {core: count}}."""
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_name("参考２")
    # header row with year labels like "14\n('02)" etc.
    hdr_r = None
    for r in range(ws.nrows):
        joined = " ".join(str(ws.cell_value(r, c)) for c in range(ws.ncols))
        if "'02)" in joined or "（2000" in joined or "平成12年" in joined:
            hdr_r = r
            break
    years = {}
    for c in range(ws.ncols):
        v = str(ws.cell_value(hdr_r, c))
        m = re.search(r"'?(\d{2})\)", v)
        if m:
            yy = int(m.group(1))
            years[c] = 2000 + yy
        elif "平成12" in v or "（2000" in v:
            years[c] = 2000
    out = {y: {k: 0.0 for k in CORE} for y in years.values()}
    for r in range(hdr_r + 1, ws.nrows):
        lab = ws.cell_value(r, 2)
        b = base(lab)
        if b not in MAP:
            continue
        core = MAP[b]
        for c, y in years.items():
            n = _num(ws.cell_value(r, c))
            if n is not None:
                out[y][core] += n
    return out


def parse_hyo4_xlsx(path):
    """表4 single-year all-specialty table (.xlsx) -> {core: count}."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sn = next(s for s in wb.sheetnames if s in ("表4", "表４"))
    ws = wb[sn]
    res = {k: 0.0 for k in CORE}
    for row in ws.iter_rows(values_only=True):
        # find label cell + first numeric to its right on same row
        cells = list(row)
        for i, v in enumerate(cells):
            b = base(v) if isinstance(v, str) else ""
            if b in MAP:
                # value: first numeric cell after i
                for w in cells[i + 1:]:
                    n = _num(w)
                    if n is not None:
                        res[MAP[b]] += n
                        break
                break
    return res


def sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def main():
    prov = []
    series = {}  # year -> {core: count}

    p18 = os.path.join(RAW, "phys_2018_toukeihyo.xls")
    s = parse_sanko2(p18)
    for y in range(2004, 2019, 2):
        if y in s:
            series[y] = s[y]
    prov.append({"years": "2000-2018 (used 2004-2018)", "sheet": "参考２",
                 "file": os.path.basename(p18), "sha256": sha256(p18)})

    for y in (2020, 2022, 2024):
        p = os.path.join(RAW, f"phys_{y}_hyo.xlsx")
        series[y] = parse_hyo4_xlsx(p)
        prov.append({"year": y, "sheet": "表４", "file": os.path.basename(p),
                     "sha256": sha256(p)})

    years = sorted(series)
    df = pd.DataFrame({"specialty": CORE})
    for y in years:
        df[y] = [int(round(series[y][k])) for k in CORE]
    out_csv = os.path.join(HERE, "physicians_by_specialty.csv")
    df.to_csv(out_csv, index=False)
    with open(os.path.join(HERE, "provenance_physicians.json"), "w") as f:
        json.dump({"source": "MHLW 医師・歯科医師・薬剤師統計",
                   "note": "biennial; broad categories matched to Supreme Court "
                           "litigation specialties; see build_physicians.py header",
                   "files": prov}, f, ensure_ascii=False, indent=2)
    print(df.to_string(index=False))
    print("\nwrote", out_csv)


if __name__ == "__main__":
    main()
